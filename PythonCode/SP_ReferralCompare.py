# -*- coding: utf-8 -*-
"""
Created on Thu Jan 25 16:53:37 2024

@author: ChristianOrtiz
"""

import os
import openpyxl
import tkinter.messagebox
import tkinter.simpledialog
from tkinter.filedialog import askdirectory, askopenfile, askopenfiles, askopenfilename
#custom
from spmodules import get_date



"""Idea
Date compared will be (yesterday). (Custom Date button)(weekend button)
Select the SP file
Select the myUnity Referral file

NEXT STEPS:
    Try to get a method in dict constructor for referral date
    have it print a message box with error, instead of adding to console.
"""

#definitions

#return a string with all the error rows.
def errorString(errorList):
    #initialize the string we'll return
    list_of_rows = str()
    #while we are looking through each List item
    for index, row_num in enumerate(errorList):
        #as long as we are not on the last index
        if index != len(errorList)-1:
            #append the row to the string
            list_of_rows += str(row_num) +', '
        else:#otherwise, add a period.
            list_of_rows += str(row_num) +'.'
    #all done. NOw return the string.
    return  list_of_rows
        
#Pop Up message to select a file(s)
tkinter.messagebox.showinfo(title="SP File Select", message="Select the SP file.")
fileName1 = askopenfilename(title='File Select',filetypes=[("Excel files", "*.xlsx"),("CSV Files", "*.csv")])
'''Continue Here - Start being able to request a date to compare'''
#tkinter.simpledialog.askstring(title='Date', prompt='Enter a Date')
currentDate = get_date()

''' uncomment once we get 1 wb imported
tkinter.messagebox.showinfo(title="SP File Select", message="Select the Referral Report file.")
fileName2 = askopenfilename(title='File Select',filetypes=[("Excel files", "*.xlsx"),("CSV Files", "*.csv")])
#file name(s) returned, with directory path.
'''



#load the first workbook
#differentiate if file is a xlsx or csv file. handle differently.
try: 
    #try to open the file as excel file
    wb1 = openpyxl.load_workbook(fileName1)
    ws1 = wb1.active
except:
    #process as csv
    pass

''' uncomment part 2
#load the second workbook
#differentiate if file is a xlsx or csv file. handle differently.
try:
    #try to open the file as excel file
    wb2 = openpyxl.load_workbook(fileName2)
    ws2 = wb1.active
except:
    #process as csv
    pass
'''

#and at this point, we have SP wb and csv wb.


#wbs to arrays
#filter the SP workbook to show only selected referral date.
#First, figure which is the SP wb.

#scan through each line to figure out which file we're looking at.

#look for the headerRow, return the number of the row if found.
#create sp array
sp_list = []

#to keep track of which row we are in
row_num=2
#a list to keep track of the number or error rows.
errors = []
#iterate through each row in the excel sp file.
for row in ws1.iter_rows(
        min_row=2, max_row=ws1.max_row,min_col=1,max_col=ws1.max_column, values_only=True):
    #if we are currently in the header row of the SP file
    #if row[0]='Referral Date' and row[1]='Branch':
        #Begin setting the dictionary keys
    current_patient = {
        "Referral Date": row[0],
        "Branch": row[1],
        "City": row[2],
        "Zip Code": row[3],
        "Patient Name": row[4],
        "Orders": row[5],
        "D1": row[6],
        "D2": row[7],
        "D3": row[8],
        "D4": row[9],
        "D5": row[10],
        "D6": row[12],
        "D7": row[12]
    }
    #format the Referral Date to MM/DD/YYYY
    
    try:   
        current_patient['Referral Date'] = current_patient['Referral Date'].strftime("%m/%d/%y")
    except Exception:
        #print('issue with row: ', row_num)
        errors.append(row_num)
    #add patient information to sp_list array/list.
    sp_list.append(current_patient)
    row_num +=1   
        
    #print(value)
    #headerRow = False
    #if value == 'Referral Date':
       # headerRow = True;
    
''' lets start fresh. may not use this code

#get the referral report list and place into an array.
referrals = list()
referrals.append(['Referral Date', 'Branch', 'City', 'Zip Code', 'Patient Name', 'Orders', 'D1'])
referralDateIndex = ''
branchIndex = ''
print('Total number of rows: '+str(ws1.max_row)+'. And total number of columns'+str(ws1.max_column))

values = [ws1.cell(row=1,column=i).value for i in range (1,ws1.max_column+1)]
print(values)

for value in ws1.iter_rows(
        min_row=1, max_row=ws1.max_row,min_col=1,max_col=ws1.max_column, values_only=True):
    
    print(value)
'''

tkinter.messagebox.showinfo(title="Results", message='Total number of rows: '+str(ws1.max_row)+'. And total number of columns: '+str(ws1.max_column)
                            +'\n Issue with row: '+ errorString(errors)
)

#for value in ws.iter_columns(min_row=1, max_row=1, min_col=1,max_col=ws.max_column, values_only=True):
#   if(this.value == 'KP Los Angeles') return ws.cell.column

